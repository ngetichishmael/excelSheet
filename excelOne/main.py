from openpyxl import *

wb = load_workbook('data.xlsx')
workingBook = Workbook()
workingSheet = workingBook.active
workingSheet.title = "Party"
count83 = 0
count100 = 0
count108 = 0

for sheet in range(3, 84):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    for row in range(4, 30):
        # print(ws['H' + str(row)].value)
        workingSheet.append([ws['H' + str(row)].value])
        count83 += 1
print(count83)

for sheet in range(85, 101):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    for row in range(4, 73):
        if ws['F' + str(row)].value is not None:
            workingSheet.append([ws['F' + str(row)].value])
            # print(ws['F' + str(row)].value)
            count100 += 1
print(count100)

for sheet in range(102, 108):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    for row in range(5, 52):
        if ws['H' + str(row)].value is not None:
            workingSheet.append([ws['H' + str(row)].value])
            # print(ws['H' + str(row)].value)
            count108 += 1
print(count108)


print(count83)
print(count100)
print(count108)

workingBook.save("PartyFinal.xlsx")
