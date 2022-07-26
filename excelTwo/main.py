from openpyxl import *

wb = load_workbook('data.xlsx')
membersWorkbook = Workbook()
senatorsWorkbook = Workbook()
governorsWorkbook = Workbook()
membersParliament = membersWorkbook.active
membersParliament.title = "Party"
senator = senatorsWorkbook.active
senator.title = "Party"
governor = governorsWorkbook.active
governor.title = "Party"
count2 = 0
count83 = 0
count100 = 0
count108 = 0

for sheet in range(2, 3):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    col_nameF = 'G'
    col_nameO = 'I'
    col_political = 'M'
    col_county = 'B'
    col_Constituency = 'E'

    for row in range(7, 16):
        # print(ws['H' + str(row)].value)
        full_name = ws[col_nameF + str(row)].value + " " + ws[col_nameO + str(row)].value
        political_party = ws[col_political + str(row)].value
        electoral_position = "MP"
        county_code = ws[col_county + str(row)].value
        constituency = ws[col_Constituency + str(row)].value
        # print(full_name, political_party, electoral_position, county_code)
        membersParliament.append([full_name, political_party, electoral_position, constituency, county_code])
        count2 += 1
print(count2)

for sheet in range(3, 84):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    col_nameF = 'E'
    col_nameO = 'F'
    col_political = 'H'
    col_county = 'B'
    col_Constituency = 'D'
    for row in range(4, 30):
        # print(ws['H' + str(row)].value)
        full_name = ws[col_nameF + str(row)].value + " " + ws[col_nameO + str(row)].value
        political_party = ws[col_political + str(row)].value
        electoral_position = "MP"
        county_code = ws[col_county + str(row)].value
        constituency = ws[col_Constituency + str(row)].value
        # print(full_name, political_party, electoral_position, county_code)
        membersParliament.append([full_name, political_party, electoral_position, constituency, county_code])
        count83 += 1
print(count83)


for sheet in range(85, 101):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    col_nameF = 'C'
    col_nameO = 'D'
    col_political = 'F'
    col_area = 'B'
    for row in range(4, 73):
        if ws['F' + str(row)].value is not None:
            # print(ws['H' + str(row)].value)
            full_name = ws[col_nameF + str(row)].value + " " + ws[col_nameO + str(row)].value
            political_party = ws[col_political + str(row)].value
            electoral_position = "Senator"
            county_code = ws[col_area + str(row)].value
            # print(full_name, political_party, electoral_position, county_code)
            senator.append([full_name, political_party, electoral_position, county_code])
            count100 += 1
print(count100)

for sheet in range(101, 102):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    col_nameF = 'E'
    col_nameO = 'G'
    col_political = 'K'
    col_county = 'C'

    for row in range(3, 16):
        # print(ws['H' + str(row)].value)
        full_name = ws[col_nameF + str(row)].value + " " + ws[col_nameO + str(row)].value
        political_party = ws[col_political + str(row)].value
        electoral_position = "Senator"
        county_code = ws[col_county + str(row)].value
        # print(full_name, political_party, electoral_position, county_code)
        senator.append([full_name, political_party, electoral_position, county_code])
        count100 += 1

print("After 100")
print(count100)


for sheet in range(101, 102):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    col_nameF = 'D'
    col_nameO = 'F'
    col_political = 'L'
    col_county = 'B'

    for row in range(20, 48):
        # print(ws['H' + str(row)].value)
        full_name = ws[col_nameF + str(row)].value + " " + ws[col_nameO + str(row)].value
        political_party = ws[col_political + str(row)].value
        electoral_position = "Governor"
        county_code = ws[col_county + str(row)].value
        # print(full_name, political_party, electoral_position, county_code)
        governor.append([full_name, political_party, electoral_position, county_code])
        count100 += 1

print("Before 102")
print(count100)
for sheet in range(102, 108):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    col_nameF = 'C'
    col_nameO = 'D'
    col_political = 'H'
    col_area = 'B'
    for row in range(5, 52):
        if ws['H' + str(row)].value is not None:
            # print(ws['H' + str(row)].value)
            full_name = ws[col_nameF + str(row)].value + " " + ws[col_nameO + str(row)].value
            political_party = ws[col_political + str(row)].value
            electoral_position = "Governor"
            county_code = ws[col_area + str(row)].value
            # print(full_name, political_party, electoral_position, county_code)
            governor.append([full_name, political_party, electoral_position, county_code])
            count108 += 1
print(count108)
for sheet in range(108, 109):
    current = 'Table ' + str(sheet)
    ws = wb[current]
    col_nameF = 'C'
    col_nameO = 'E'
    col_political = 'I'
    col_county = 'B'

    for row in range(3, 4):
        # print(ws['H' + str(row)].value)
        full_name = ws[col_nameF + str(row)].value + " " + ws[col_nameO + str(row)].value
        political_party = ws[col_political + str(row)].value
        electoral_position = "Governor"
        county_code = ws[col_county + str(row)].value
        # print(full_name, political_party, electoral_position, county_code)
        governor.append([full_name, political_party, electoral_position, county_code])
        count108 += 1

print("After 108")
print(count108)
senatorsWorkbook.save("Senators.xlsx")
governorsWorkbook.save("Governors.xlsx")
membersWorkbook.save("Members.xlsx")
